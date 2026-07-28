import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def determinar_pedir(row):
    """
    Determina si se debe pedir un producto basándose en stock y consumo a 6 meses.
    """
    stock_disp = row["STOCK DISP"]
    stock_p_recibir = row["STOCK.P.RECIBIR"]
    cantidad = row["CANTIDAD"]
    
    # Si existe columna SITUACION, usarla como referencia rápida
    if "SITUACION" in row.index:
        situacion = row["SITUACION"]
        if situacion == "MATERIAL NO PEDIR":
            return "NO PEDIR"
        if situacion == "BAJO PEDIDO":
            return "REVISAR SI PEDIR"
    
    # Validaciones de datos
    if pd.isna(stock_disp) or pd.isna(cantidad) or pd.isna(stock_p_recibir):
        return "DATOS INCOMPLETOS"
    
    if cantidad <= 0:
        return "DATOS INCOMPLETOS"
    
    if stock_disp < 0 or stock_p_recibir < 0:
        return "DATOS INCOMPLETOS"
    
    # Calcular umbrales basados en consumo a 6 meses
    umbral_pedir = cantidad * 0.5
    umbral_revisar = cantidad * 0.6
    
    # Stock total (disponible + en tránsito)
    stock_total = stock_disp + stock_p_recibir
    
    # Decisión
    if stock_total >= umbral_revisar:
        return "NO PEDIR"
    elif stock_total >= umbral_pedir:
        return "REVISAR SI PEDIR"
    else:
        return "PEDIR"


def procesar_excel(stock_path: str):
    """
    Procesa archivo de stock y determina qué productos pedir.
    
    Args:
        stock_path: Ruta del archivo de stock Excel
    
    Returns:
        Tupla (ruta_salida, resumen_resultados)
    """
    # Leer datos de stock
    resultado = pd.read_excel(stock_path)
        
    # Aplicar lógica de decisión
    resultado["pedir"] = resultado.apply(determinar_pedir, axis=1)
    
    # Ordenar por categoría de decisión
    orden_custom = {
        "PEDIR": 1,
        "REVISAR SI PEDIR": 2,
        "NO PEDIR": 3,
        "DATOS INCOMPLETOS": 4
    }
    resultado["_orden"] = resultado["pedir"].map(orden_custom)
    resultado = resultado.sort_values("_orden").drop("_orden", axis=1)
    resultado = resultado.reset_index(drop=True)
    
    # Guardar resultado
    salida = Path(stock_path).with_name(
        Path(stock_path).stem + "_Resultado.xlsx"
    )
    resultado.to_excel(salida, index=False)
    
    # Aplicar colores con openpyxl
    _aplicar_colores(salida)
    
    # Generar resumen
    conteo = resultado["pedir"].value_counts()
    resumen = {
        "PEDIR": conteo.get("PEDIR", 0),
        "REVISAR SI PEDIR": conteo.get("REVISAR SI PEDIR", 0),
        "NO PEDIR": conteo.get("NO PEDIR", 0),
        "DATOS INCOMPLETOS": conteo.get("DATOS INCOMPLETOS", 0),
    }
    
    return str(salida), resumen


def _aplicar_colores(excel_path: str):
    """
    Aplica colores SOLO a la columna 'pedir' según su valor.
    Rojo: pedir, Verde: no pedir, Naranja: revisar si pedir, Gris: datos incompletos
    """
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # Definir colores
    colores = {
        "PEDIR": "FF0000",  # Rojo
        "NO PEDIR": "00B050",  # Verde
        "REVISAR SI PEDIR": "FFA500",  # Naranja
        "DATOS INCOMPLETOS": "D3D3D3"  # Gris claro
    }
    
    # Encontrar índice de la columna "pedir"
    columna_pedir = None
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value == "pedir":
            columna_pedir = col_idx
            break
    
    if columna_pedir is None:
        return
    
    # Aplicar colores SOLO a la columna "pedir"
    for row_idx in range(2, ws.max_row + 1):
        valor_pedir = ws.cell(row_idx, columna_pedir).value
        color = colores.get(valor_pedir)
        
        if color:
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws.cell(row_idx, columna_pedir).fill = fill
    
    wb.save(excel_path)